import io
import logging

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils.exceptions import SheetTitleException

from upload_tools import upload_file
from .helpers import add_table_to_sheet
from .parser import (
    walk_markdown_lines,
    collect_table_positions,
    SheetEvent,
    HeaderEvent,
    TableEvent,
    DEFAULT_SHEET_NAME,
    _sanitize_sheet_name,
)

logger = logging.getLogger(__name__)

# ── Constants ──────────────────────────────────────────────────────────────────
# Header font styles by level
HEADER_FONTS = {
    1: Font(size=16, bold=True, color="2F5597"),
    2: Font(size=14, bold=True, color="4472C4"),
}
HEADER_FONT_DEFAULT = Font(size=12, bold=True)


def markdown_to_excel(markdown_content: str, file_name: str | None = None, auto_filter: bool = False) -> str:
    """Convert Markdown to Excel workbook (focused on tables and headers).

    Always starts from an empty Workbook (no templates).
    Supports multiple sheets via '## Sheet: Name' headings.
    Supports cross-sheet references via ``SheetName!T1.B[0]`` syntax.

    Args:
        markdown_content: Markdown string with tables.
        file_name: Optional custom filename (without extension).
        auto_filter: If True, apply Excel auto-filter to each table.

    Raises:
        RuntimeError: If the markdown contains no tables or conversion fails.
    """
    logger.info("Starting markdown_to_excel conversion")

    # ── Input validation ──
    if not markdown_content or not markdown_content.strip():
        raise RuntimeError("Cannot create Excel workbook: markdown content is empty")

    # Split content into lines and parse into events (single shared state machine)
    lines: list[str] = markdown_content.split('\n')
    events = walk_markdown_lines(lines)

    # Build table position map from events (used for cross-sheet formula resolution)
    all_sheet_table_positions = collect_table_positions(events)
    logger.debug("Table positions (all sheets): %s", all_sheet_table_positions)

    # ── Build the actual workbook from events ──
    wb = Workbook()
    ws = wb.active
    ws.title = _sanitize_sheet_name(DEFAULT_SHEET_NAME)

    # Per-sheet state for formula resolution
    table_positions: dict[str, int] = {}

    # Counters for summary
    headers_count = 0
    tables_count = 0

    try:
        for event in events:
            if isinstance(event, SheetEvent):
                if event.is_rename:
                    try:
                        ws.title = event.sheet_name
                    except (SheetTitleException, ValueError) as exc:
                        logger.warning(
                            "Could not rename worksheet to '%s': %s — using default",
                            event.sheet_name, exc,
                        )
                else:
                    try:
                        ws = wb.create_sheet(title=event.sheet_name)
                    except (SheetTitleException, ValueError) as exc:
                        logger.warning(
                            "Invalid sheet name '%s': %s — using fallback",
                            event.sheet_name, exc,
                        )
                        ws = wb.create_sheet()
                    table_positions = {}

            elif isinstance(event, HeaderEvent):
                cell = ws.cell(row=event.row, column=1)
                cell.value = event.text
                cell.font = HEADER_FONTS.get(event.level, HEADER_FONT_DEFAULT)
                headers_count += 1
                logger.debug("Header (level %d) at row %d: %s", event.level, event.row, event.text)

            elif isinstance(event, TableEvent):
                # Record this table's position for local formula resolution
                table_positions[event.table_key] = event.start_row

                # Write table to worksheet
                add_table_to_sheet(
                    event.table_data, ws, event.start_row, table_positions,
                    all_sheet_table_positions=all_sheet_table_positions,
                    auto_filter=auto_filter,
                    table_index=tables_count,
                    directives=event.directives,
                )

                # Handle freeze directive — freeze below header row of this table
                if 'freeze' in event.directives:
                    ws.freeze_panes = f"A{event.start_row + 1}"

                tables_count += 1
                logger.debug(
                    "Added table #%d (%s) with %d data rows on sheet '%s'",
                    tables_count, event.table_key, len(event.table_data) - 1, event.sheet_name,
                )

    except Exception as e:
        logger.error("Error generating Excel workbook: %s", str(e), exc_info=True)
        raise RuntimeError(f"Error generating Excel workbook: {e}") from e

    # ── Validation: ensure at least one table was created ──
    if tables_count == 0:
        raise RuntimeError(
            "Cannot create Excel workbook: no valid markdown tables found in the input. "
            "Tables must use pipe syntax (| col1 | col2 |) with a separator row (|---|---|)."
        )

    # Save workbook to BytesIO and upload via existing helper
    file_object = io.BytesIO()
    try:
        logger.info("Saving Excel workbook to memory buffer (headers=%d, tables=%d)", headers_count, tables_count)
        wb.save(file_object)
        file_object.seek(0)
        result = upload_file(file_object, "xlsx", filename=file_name)
        logger.info("Excel upload completed successfully")
        return result
    except Exception as e:
        logger.error("Error saving/uploading Excel workbook: %s", str(e), exc_info=True)
        raise RuntimeError(f"Error saving/uploading Excel workbook: {e}") from e
    finally:
        file_object.close()
