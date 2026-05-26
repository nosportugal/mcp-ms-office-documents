import re
import logging
from dataclasses import dataclass
from datetime import datetime

from dateutil import parser as dateutil_parser
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

logger = logging.getLogger(__name__)

# ── Layout Constants ──────────────────────────────────────────────────────────
TABLE_BOTTOM_SPACING = 2
MIN_COLUMN_WIDTH = 12
MAX_COLUMN_WIDTH = 25
COLUMN_WIDTH_PADDING = 2

# Date formats to try before falling back to dateutil auto-detection.
# Order matters — more specific/common formats first.
# Each entry: (strptime_format, excel_number_format)
DATE_FORMATS: list[tuple[str, str]] = [
    # ISO
    ("%Y-%m-%d", "YYYY-MM-DD"),
    ("%Y-%m-%dT%H:%M:%S", "YYYY-MM-DD HH:MM:SS"),
    ("%Y-%m-%dT%H:%M", "YYYY-MM-DD HH:MM"),
    # European (day first)
    ("%d.%m.%Y", "DD.MM.YYYY"),
    ("%d/%m/%Y", "DD/MM/YYYY"),
    ("%d-%m-%Y", "DD-MM-YYYY"),
    ("%d. %m. %Y", "DD. MM. YYYY"),
    # US (month first)
    ("%m/%d/%Y", "MM/DD/YYYY"),
    # With time
    ("%d.%m.%Y %H:%M", "DD.MM.YYYY HH:MM"),
    ("%d.%m.%Y %H:%M:%S", "DD.MM.YYYY HH:MM:SS"),
    ("%m/%d/%Y %H:%M", "MM/DD/YYYY HH:MM"),
    # Short year
    ("%d.%m.%y", "DD.MM.YY"),
    ("%d/%m/%y", "DD/MM/YY"),
    ("%m/%d/%y", "MM/DD/YY"),
    # Named months
    ("%d %b %Y", "DD MMM YYYY"),
    ("%d %B %Y", "DD MMMM YYYY"),
    ("%b %d, %Y", "MMM DD, YYYY"),
    ("%B %d, %Y", "MMMM DD, YYYY"),
]

# Minimum length to even attempt date parsing (avoids matching plain numbers)
_MIN_DATE_LENGTH = 6
# Regex to quickly reject values that clearly can't be dates
_DATE_CANDIDATE_RE = re.compile(r'^\d{1,4}[\.\-/]|^\d{1,2}\s+\w|^\w+\s+\d')


def _try_parse_date(value: str) -> tuple[datetime, str] | None:
    """Attempt to parse a string as a date/datetime.

    Tries explicit formats first (fast, unambiguous), then falls back to
    dateutil for natural language dates.

    Returns (datetime_obj, excel_number_format) or None.
    """
    if len(value) < _MIN_DATE_LENGTH:
        return None
    if not _DATE_CANDIDATE_RE.match(value):
        return None

    # Try explicit formats first (deterministic, no ambiguity)
    for fmt, xl_fmt in DATE_FORMATS:
        try:
            dt = datetime.strptime(value, fmt)
            return dt, xl_fmt
        except ValueError:
            continue

    # Fallback to dateutil (handles many international/natural formats)
    try:
        dt = dateutil_parser.parse(value, dayfirst=True, fuzzy=False)
        # Only accept if the string is sufficiently "date-like" —
        # dateutil can parse things like "1" or "March" alone which we don't want
        if dt and len(value) >= 8:
            # Determine appropriate format based on whether time is present
            if dt.hour or dt.minute or dt.second:
                return dt, "YYYY-MM-DD HH:MM:SS"
            return dt, "YYYY-MM-DD"
    except (ValueError, TypeError, OverflowError):
        pass

    return None


def _is_separator_row(line: str) -> bool:
    """Check if a table line is a markdown separator row (e.g. |---|:---:|---:|).

    Only returns True if ALL cells in the row match the separator pattern,
    preventing false positives from data cells that happen to contain '---'.
    """
    cells = [c.strip() for c in line.split('|')[1:-1]]
    if not cells:
        return False
    return all(re.match(r'^:?-{3,}:?$', c) for c in cells)


def _parse_column_alignments(separator_line: str) -> list[str | None]:
    """Extract column alignments from a markdown separator row.

    Returns a list of alignment strings ('left', 'center', 'right') or None per column.
    This is the same logic used by docx_tools but returns generic strings
    instead of Word-specific enums.
    """
    cells = [c.strip() for c in separator_line.split('|')[1:-1]]
    alignments: list[str | None] = []
    for cell in cells:
        cell = cell.strip()
        if cell.startswith(':') and cell.endswith(':'):
            alignments.append('center')
        elif cell.endswith(':'):
            alignments.append('right')
        elif cell.startswith(':'):
            alignments.append('left')
        else:
            alignments.append(None)  # auto — will use heuristic
    return alignments


def parse_table(lines: list[str], start_idx: int) -> tuple[list[list[str]] | None, int]:
    """Parse markdown table and return (table_data, next_index).

    Also extracts column alignments from the separator row and attaches them
    as the 'col_alignments' attribute on the returned TableData instance.
    """
    table_lines: list[str] = []
    i = start_idx

    # Find all consecutive table lines (allow missing trailing pipe)
    while i < len(lines):
        line = lines[i].strip()
        if line.startswith('|'):
            # Normalize: ensure trailing pipe for consistent splitting
            if not line.endswith('|'):
                line = line + '|'
            table_lines.append(line)
            i += 1
        else:
            break

    if len(table_lines) < 2:  # Need at least header and separator
        return None, i if i > start_idx else start_idx + 1

    # Parse table data, extracting alignment from separator row
    table_data: list[list[str]] = []
    col_alignments: list[str | None] = []
    for line in table_lines:
        if _is_separator_row(line):
            col_alignments = _parse_column_alignments(line)
            continue
        cells = [cell.strip() for cell in line.split('|')[1:-1]]
        table_data.append(cells)

    # Attach alignment info to the table_data list
    table_data_with_align = TableData(table_data, col_alignments)
    return table_data_with_align, i


class TableData(list):
    """A list subclass that carries column alignment metadata."""

    def __init__(self, data: list[list[str]], col_alignments: list[str | None] | None = None):
        super().__init__(data)
        self.col_alignments: list[str | None] = col_alignments or []


# ── Cell Resolution ───────────────────────────────────────────────────────────

@dataclass
class CellResult:
    """Resolved cell metadata — all information needed to write a cell to Excel."""
    value: str | int | float | datetime  # The cleaned value to write
    is_formula: bool = False
    is_percent: bool = False
    is_date: bool = False
    date_format: str = ""  # Excel number format for dates (e.g. "YYYY-MM-DD")
    bold: bool = False
    italic: bool = False
    monospace: bool = False

    @property
    def formatting_info(self) -> dict[str, bool]:
        """Legacy-compatible formatting dict for apply_cell_formatting()."""
        return {'bold': self.bold, 'italic': self.italic, 'monospace': self.monospace}




def _strip_markdown_formatting(raw_text: str) -> tuple[str, dict[str, bool]]:
    """Strip inline markdown formatting markers from a cell value.

    Returns (clean_text, formatting_dict) where formatting_dict has
    'bold', 'italic', 'monospace' keys.
    """
    clean_text = raw_text.strip()
    formatting = {'bold': False, 'italic': False, 'monospace': False}

    if clean_text.startswith('**') and clean_text.endswith('**') and len(clean_text) > 4:
        clean_text = clean_text[2:-2]
        formatting['bold'] = True
    elif clean_text.startswith('*') and clean_text.endswith('*') and len(clean_text) > 2:
        clean_text = clean_text[1:-1]
        formatting['italic'] = True
    elif clean_text.startswith('`') and clean_text.endswith('`') and len(clean_text) > 2:
        clean_text = clean_text[1:-1]
        formatting['monospace'] = True

    return clean_text, formatting


def resolve_cell(raw_text: str) -> CellResult:
    """Parse a raw markdown cell string into a fully resolved CellResult.

    Combines formatting detection, formula detection, and type conversion
    in a single pass — the unified replacement for the former three-function pipeline
    of parse_cell_formatting → detect_formula_pattern → format_cell_value.
    """
    # Step 1: Strip markdown formatting markers
    clean_text, formatting = _strip_markdown_formatting(raw_text)
    bold = formatting['bold']
    italic = formatting['italic']
    monospace = formatting['monospace']

    # Step 2: Check if it's an explicit formula (= prefix)
    if clean_text.startswith('='):
        return CellResult(
            value=clean_text, is_formula=True,
            bold=bold, italic=italic, monospace=monospace,
        )

    # Step 3: Detect percent and convert to number
    is_percent = clean_text.endswith('%')
    if is_percent:
        try:
            numeric_val = float(clean_text[:-1]) / 100
            return CellResult(
                value=numeric_val, is_percent=True,
                bold=bold, italic=italic, monospace=monospace,
            )
        except ValueError:
            pass  # Not a valid percent number — fall through

    # Step 5: Try numeric conversion
    try:
        numeric_val = float(clean_text)
        return CellResult(
            value=numeric_val,
            bold=bold, italic=italic, monospace=monospace,
        )
    except ValueError:
        pass

    # Step 6: Try date detection (after numeric, so "2024" isn't parsed as a date)
    date_result = _try_parse_date(clean_text)
    if date_result:
        dt, xl_fmt = date_result
        return CellResult(
            value=dt, is_date=True, date_format=xl_fmt,
            bold=bold, italic=italic, monospace=monospace,
        )

    # Step 7: Plain text
    return CellResult(
        value=clean_text,
        bold=bold, italic=italic, monospace=monospace,
    )


def apply_cell_formatting(cell, formatting_info: dict[str, bool]) -> None:
    """Apply formatting information to an Excel cell."""
    current_font = cell.font
    if formatting_info['bold']:
        cell.font = Font(bold=True, color=current_font.color, size=current_font.size)
    elif formatting_info['italic']:
        cell.font = Font(italic=True, color=current_font.color, size=current_font.size)
    elif formatting_info['monospace']:
        cell.font = Font(name='Courier New', color=current_font.color, size=current_font.size)


# ── Formula Reference Resolution ─────────────────────────────────────────────

def _quote_sheet_name(name: str) -> str:
    """Return the sheet name quoted for Excel if it contains spaces or special chars."""
    if re.search(r"[^A-Za-z0-9_]", name):
        return f"'{name}'"
    return name


def _resolve_row(positions: dict[str, int], table_num: int, offset: int, fallback_row: int) -> int:
    """Resolve a table-relative row reference to an absolute Excel row number.

    Args:
        positions: Table positions dict ({"T1": start_row, ...}) for the target sheet.
        table_num: Table number (1-based).
        offset: Row offset within the table (0 = first data row).
        fallback_row: Row to use if the table isn't found in positions.

    Returns:
        The absolute Excel row number.
    """
    key = f"T{table_num}"
    base = positions.get(key)
    if base is not None:
        return base + 1 + offset  # +1 to skip header row
    return fallback_row + offset


def _make_cell_ref(column: str, row: int, sheet: str | None = None) -> str:
    """Build a cell reference string, optionally with a quoted sheet prefix."""
    if sheet:
        return f"{_quote_sheet_name(sheet)}!{column}{row}"
    return f"{column}{row}"


def adjust_formula_references(
    formula: str,
    current_excel_row: int,
    table_positions: dict[str, int] | None = None,
    all_sheet_table_positions: dict[str, dict[str, int]] | None = None,
) -> str:
    """Convert row-relative references [offset] and table references T1.B[1] to actual Excel row numbers.

    Also resolves cross-sheet references like ``SheetName!T1.B[0]`` → ``'SheetName'!B2``.
    """
    if not formula.startswith('='):
        return formula

    if table_positions is None:
        table_positions = {}
    if all_sheet_table_positions is None:
        all_sheet_table_positions = {}

    logger.debug("Resolving formula: %s (current_row=%d)", formula, current_excel_row)

    try:
        # ── Cross-sheet references (must be resolved BEFORE local patterns) ──

        # Cross-sheet function: SheetName!T1.SUM(B[0]:E[0])
        cs_func_pattern = r"([\w\s.]+)!T(\d+)\.(SUM|AVERAGE|MAX|MIN)\(([A-Z]+)\[([+-]?\d+)\]:([A-Z]+)\[([+-]?\d+)\]\)"

        def _replace_cs_func(match):
            sheet = match.group(1).strip()
            table_num = int(match.group(2))
            func_name = match.group(3)
            start_col = match.group(4)
            start_offset = int(match.group(5))
            end_col = match.group(6)
            end_offset = int(match.group(7))
            pos = all_sheet_table_positions.get(sheet, {})
            sr = _resolve_row(pos, table_num, start_offset, current_excel_row)
            er = _resolve_row(pos, table_num, end_offset, current_excel_row)
            qs = _quote_sheet_name(sheet)
            result = f"{func_name}({qs}!{start_col}{sr}:{qs}!{end_col}{er})"
            logger.debug("  Cross-sheet func: %s → %s", match.group(0), result)
            return result

        formula = re.sub(cs_func_pattern, _replace_cs_func, formula)

        # Cross-sheet range: SheetName!T1.B[0]:T1.E[0]
        cs_range_pattern = r"([\w\s.]+)!T(\d+)\.([A-Z]+)\[([+-]?\d+)\]:T(\d+)\.([A-Z]+)\[([+-]?\d+)\]"

        def _replace_cs_range(match):
            sheet = match.group(1).strip()
            st_num = int(match.group(2))
            start_col = match.group(3)
            start_offset = int(match.group(4))
            et_num = int(match.group(5))
            end_col = match.group(6)
            end_offset = int(match.group(7))
            pos = all_sheet_table_positions.get(sheet, {})
            sr = _resolve_row(pos, st_num, start_offset, current_excel_row)
            er = _resolve_row(pos, et_num, end_offset, current_excel_row)
            qs = _quote_sheet_name(sheet)
            result = f"{qs}!{start_col}{sr}:{end_col}{er}"
            logger.debug("  Cross-sheet range: %s → %s", match.group(0), result)
            return result

        formula = re.sub(cs_range_pattern, _replace_cs_range, formula)

        # Cross-sheet single cell: SheetName!T1.B[0]
        cs_cell_pattern = r"([\w\s.]+)!T(\d+)\.([A-Z]+)\[([+-]?\d+)\]"

        def _replace_cs_cell(match):
            sheet = match.group(1).strip()
            table_num = int(match.group(2))
            column = match.group(3)
            offset = int(match.group(4))
            pos = all_sheet_table_positions.get(sheet, {})
            actual_row = _resolve_row(pos, table_num, offset, current_excel_row)
            result = _make_cell_ref(column, actual_row, sheet)
            logger.debug("  Cross-sheet cell: %s → %s", match.group(0), result)
            return result

        formula = re.sub(cs_cell_pattern, _replace_cs_cell, formula)

        # ── Local (same-sheet) references ──
        # NOTE: Range and function patterns must be processed BEFORE single-cell
        # to prevent the single-cell regex from consuming parts of range expressions.

        # Table range references e.g. T1.B[0]:T1.E[0]
        table_range_pattern = r'T(\d+)\.([A-Z]+)\[([+-]?\d+)\]:T(\d+)\.([A-Z]+)\[([+-]?\d+)\]'

        def replace_table_range(match):
            start_table_num = int(match.group(1))
            start_col = match.group(2)
            start_offset = int(match.group(3))
            end_table_num = int(match.group(4))
            end_col = match.group(5)
            end_offset = int(match.group(6))
            start_row = _resolve_row(table_positions, start_table_num, start_offset, current_excel_row)
            end_row = _resolve_row(table_positions, end_table_num, end_offset, current_excel_row)
            return f"{start_col}{start_row}:{end_col}{end_row}"

        adjusted = re.sub(table_range_pattern, replace_table_range, formula)

        # Simplified function over table range e.g. T1.SUM(B[0]:E[0])
        table_func_pattern = r'T(\d+)\.(SUM|AVERAGE|MAX|MIN)\(([A-Z]+)\[([+-]?\d+)\]:([A-Z]+)\[([+-]?\d+)\]\)'

        def replace_table_function(match):
            table_num = int(match.group(1))
            func_name = match.group(2)
            start_col = match.group(3)
            start_offset = int(match.group(4))
            end_col = match.group(5)
            end_offset = int(match.group(6))
            start_row = _resolve_row(table_positions, table_num, start_offset, current_excel_row)
            end_row = _resolve_row(table_positions, table_num, end_offset, current_excel_row)
            return f"{func_name}({start_col}{start_row}:{end_col}{end_row})"

        adjusted = re.sub(table_func_pattern, replace_table_function, adjusted)

        # Table cell references e.g. T1.B[1] (AFTER range patterns)
        table_pattern = r'T(\d+)\.([A-Z]+)\[([+-]?\d+)\]'

        def replace_table_reference(match):
            table_num = int(match.group(1))
            column = match.group(2)
            offset = int(match.group(3))
            actual_row = _resolve_row(table_positions, table_num, offset, current_excel_row)
            result = f"{column}{actual_row}"
            logger.debug("  Local table ref: %s → %s", match.group(0), result)
            return result

        adjusted = re.sub(table_pattern, replace_table_reference, adjusted)

        # Determine current table start for relative references
        current_table_start = None
        for table_key, table_start_row in table_positions.items():
            if table_start_row <= current_excel_row:
                current_table_start = table_start_row

        # Row-relative range e.g. B[0]:E[0] (BEFORE single-cell relative)
        range_pattern = r'([A-Z]+)\[([+-]?\d+)\]:([A-Z]+)\[([+-]?\d+)\]'

        def replace_range(match):
            start_col = match.group(1)
            start_offset = int(match.group(2))
            end_col = match.group(3)
            end_offset = int(match.group(4))
            if current_table_start is not None:
                start_row = current_table_start + 1 + start_offset
                end_row = current_table_start + 1 + end_offset
            else:
                start_row = current_excel_row + start_offset
                end_row = current_excel_row + end_offset
            return f"{start_col}{start_row}:{end_col}{end_row}"

        adjusted = re.sub(range_pattern, replace_range, adjusted)

        # Handle row-relative references e.g. B[0] (AFTER range pattern)
        rel_pattern = r'([A-Z]+)\[([+-]?\d+)\]'

        def replace_rel(match):
            column = match.group(1)
            offset = int(match.group(2))
            if current_table_start is not None:
                actual_row = current_table_start + 1 + offset
            else:
                actual_row = current_excel_row + offset
            result = f"{column}{actual_row}"
            logger.debug("  Relative ref: %s → %s", match.group(0), result)
            return result

        adjusted = re.sub(rel_pattern, replace_rel, adjusted)

        logger.debug("  Resolved formula: %s → %s", formula, adjusted)
        return adjusted

    except Exception as e:
        logger.warning("Failed to adjust formula references for '%s': %s", formula, e)
        return formula


# ── Directive Helpers ──────────────────────────────────────────────────────────

# Currency symbols → Excel format string
_CURRENCY_FORMATS = {
    '$': '$#,##0.00',
    '€': '#,##0.00 €',
    '£': '£#,##0.00',
    '¥': '¥#,##0',
    'Kč': '#,##0.00 "Kč"',
    'zł': '#,##0.00 "zł"',
    'kr': '#,##0.00 "kr"',
    'CHF': '"CHF" #,##0.00',
    'R$': '"R$" #,##0.00',
    '₹': '₹#,##0.00',
}


def _parse_types_directive(value: str) -> list[str | None]:
    """Parse a types directive value like 'text, currency:$, date, bool, number'.

    Returns a list of type specs (or None for unspecified columns).
    """
    if not value:
        return []
    return [t.strip() or None for t in value.split(',')]


def _apply_column_type(cell, raw_text: str, type_spec: str | None) -> bool:
    """Apply column type coercion to a cell based on directive.

    Returns True if type was applied (caller should skip default processing),
    False if default processing should continue.
    """
    if not type_spec:
        return False

    clean = raw_text.strip()
    type_lower = type_spec.lower()

    # text — force string, no conversion
    if type_lower == 'text':
        cell.value = clean
        return True

    # bool — map common boolean strings to Excel boolean
    if type_lower == 'bool':
        lower_val = clean.lower()
        if lower_val in ('true', 'yes', '1', 'on'):
            cell.value = True
        elif lower_val in ('false', 'no', '0', 'off'):
            cell.value = False
        else:
            cell.value = clean  # Unrecognized → keep as text
        return True

    # currency:<symbol> — strip symbol and thousands separators, store as number
    if type_lower.startswith('currency'):
        symbol = type_spec.split(':', 1)[1].strip() if ':' in type_spec else '$'
        if not symbol:
            symbol = '$'  # Default if directive is 'currency:' with no symbol
        # Strip the currency symbol and common thousand separators
        numeric_str = clean.replace(symbol, '').replace(' ', '').strip()
        # Handle both comma-as-thousands (1,234.56) and dot-as-thousands (1.234,56)
        if ',' in numeric_str and '.' in numeric_str:
            # Determine which is the decimal separator (last one wins)
            last_comma = numeric_str.rfind(',')
            last_dot = numeric_str.rfind('.')
            if last_comma > last_dot:
                # European: 1.234,56
                numeric_str = numeric_str.replace('.', '').replace(',', '.')
            else:
                # English: 1,234.56
                numeric_str = numeric_str.replace(',', '')
        elif ',' in numeric_str and '.' not in numeric_str:
            # Could be thousands (1,234) or decimal (1,5) — assume thousands if >3 digits after comma
            parts = numeric_str.split(',')
            if len(parts[-1]) == 3:
                numeric_str = numeric_str.replace(',', '')
            else:
                numeric_str = numeric_str.replace(',', '.')
        try:
            cell.value = float(numeric_str)
            cell.number_format = _CURRENCY_FORMATS.get(symbol, f'#,##0.00 "{symbol}"')
        except ValueError:
            cell.value = clean  # Can't parse → keep as text
        return True

    # number or number:<format> — parse as number with optional format
    if type_lower.startswith('number'):
        fmt = type_spec.split(':', 1)[1].strip() if ':' in type_spec else None
        numeric_str = clean.replace(',', '').replace(' ', '')
        try:
            cell.value = float(numeric_str)
            if fmt:
                cell.number_format = fmt
            elif cell.value >= 1000:
                cell.number_format = '#,##0'
        except ValueError:
            cell.value = clean
        return True

    # date or date:<format> — parse with dateutil, apply format
    if type_lower.startswith('date'):
        fmt = type_spec.split(':', 1)[1].strip() if ':' in type_spec else None
        result = _try_parse_date(clean)
        if result:
            dt, default_fmt = result
            cell.value = dt
            cell.number_format = fmt or default_fmt
        else:
            cell.value = clean
        return True

    # percent — parse as percent
    if type_lower == 'percent':
        numeric_str = clean.rstrip('%').strip()
        try:
            cell.value = float(numeric_str) / 100
            cell.number_format = '0%'
        except ValueError:
            cell.value = clean
        return True

    return False


# ── Table Rendering ───────────────────────────────────────────────────────────

def add_table_to_sheet(
    table_data: list[list[str]],
    worksheet,
    start_row: int,
    table_positions: dict[str, int] | None = None,
    all_sheet_table_positions: dict[str, dict[str, int]] | None = None,
    auto_filter: bool = False,
    table_index: int = 0,
    directives: dict[str, str] | None = None,
) -> int:
    """Add table data to Excel worksheet with proper formatting and formula support."""
    if not table_data:
        return start_row

    directives = directives or {}

    # Parse column type hints from <!-- types: text, currency:$, date, bool --> directive
    col_types: list[str | None] = _parse_types_directive(directives.get('types', ''))

    # Extract column alignments if available (from TableData subclass)
    col_alignments: list[str | None] = []
    if hasattr(table_data, 'col_alignments'):
        col_alignments = table_data.col_alignments

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    formula_fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Fill cells
    for row_idx, row_data in enumerate(table_data):
        current_excel_row = start_row + row_idx
        for col_idx, cell_text in enumerate(row_data):
            try:
                cell = worksheet.cell(row=current_excel_row, column=col_idx + 1)

                # If column type directive applies (data rows only), use it
                col_type = col_types[col_idx] if col_idx < len(col_types) else None
                if row_idx > 0 and col_type:
                    # Strip markdown formatting before type coercion
                    clean_text, fmt_info = _strip_markdown_formatting(cell_text)
                    if _apply_column_type(cell, clean_text, col_type):
                        # Type directive handled the cell value — apply formatting, border, alignment
                        apply_cell_formatting(cell, fmt_info)
                        cell.border = border
                        explicit_align = col_alignments[col_idx] if col_idx < len(col_alignments) else None
                        if explicit_align:
                            cell.alignment = Alignment(horizontal=explicit_align)
                        elif isinstance(cell.value, bool):
                            cell.alignment = Alignment(horizontal='center')
                        elif isinstance(cell.value, (int, float, datetime)):
                            cell.alignment = Alignment(horizontal='right')
                        else:
                            cell.alignment = Alignment(horizontal='left')
                        continue

                resolved = resolve_cell(cell_text)

                if resolved.is_formula:
                    adjusted_formula = adjust_formula_references(
                        resolved.value, current_excel_row, table_positions, all_sheet_table_positions
                    )
                    cell.value = adjusted_formula
                    cell.fill = formula_fill
                else:
                    # Header row must remain as strings — Excel Tables require
                    # string headers; numeric-looking headers (e.g. "2024") must
                    # not be converted to numbers.
                    if row_idx == 0:
                        # Use the original stripped text for headers to avoid
                        # artifacts like "2024.0" from float conversion
                        clean_header, _ = _strip_markdown_formatting(cell_text)
                        cell.value = clean_header
                    else:
                        cell.value = resolved.value

                # Apply inline formatting (bold/italic/monospace) — skip for header row
                # since header styling will override it immediately below
                if row_idx > 0:
                    apply_cell_formatting(cell, resolved.formatting_info)
                cell.border = border

                # Alignment — use explicit column alignment from separator if available,
                # otherwise fall back to heuristic
                explicit_align = col_alignments[col_idx] if col_idx < len(col_alignments) else None
                if row_idx == 0:
                    cell.alignment = Alignment(horizontal='center')
                elif explicit_align:
                    cell.alignment = Alignment(horizontal=explicit_align)
                elif isinstance(cell.value, (int, float, datetime)) or (isinstance(cell.value, str) and cell.value.startswith('=')):
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')

                # Header row styling (overrides inline formatting)
                if row_idx == 0:
                    cell.font = header_font
                    cell.fill = header_fill
                elif isinstance(cell.value, (int, float)) and cell.value >= 1000:
                    cell.number_format = '#,##0'

                # Apply percentage number format
                if resolved.is_percent and isinstance(cell.value, (int, float)):
                    cell.number_format = '0%'

                # Apply date number format
                if resolved.is_date and resolved.date_format:
                    cell.number_format = resolved.date_format
            except Exception as e:
                logger.warning("Error processing cell [row=%d, col=%d]: %s", current_excel_row, col_idx + 1, e)

    # Column widths — based on clean text length (not raw markdown with formatting markers)
    # When type directives are active, estimate display width from the type spec.
    FORMULA_WIDTH_CAP = 12  # Formulas display as numbers, cap their width contribution
    for col_idx in range(len(table_data[0]) if table_data else 0):
        column_letter = get_column_letter(col_idx + 1)
        col_type = col_types[col_idx] if col_idx < len(col_types) else None
        max_length = 0
        for row_idx, row in enumerate(table_data):
            if col_idx < len(row):
                # For data rows with a type directive, estimate from the directive
                if row_idx > 0 and col_type:
                    type_lower = col_type.lower()
                    if type_lower == 'bool':
                        length = 5  # "FALSE" is longest
                    elif type_lower.startswith('currency'):
                        # Symbol + number — use raw text length as decent estimate
                        length = len(row[col_idx].strip())
                    elif type_lower.startswith('date'):
                        fmt = col_type.split(':', 1)[1].strip() if ':' in col_type else "YYYY-MM-DD"
                        length = len(fmt)
                    elif type_lower == 'percent':
                        length = 6  # e.g. "85.0%"
                    else:
                        length = len(row[col_idx].strip())
                else:
                    resolved = resolve_cell(row[col_idx])
                    if resolved.is_formula:
                        length = FORMULA_WIDTH_CAP
                    elif resolved.is_date:
                        length = len(resolved.date_format)
                    else:
                        length = len(str(resolved.value))
                max_length = max(max_length, length)
        adjusted_width = min(max(max_length + COLUMN_WIDTH_PADDING, MIN_COLUMN_WIDTH), MAX_COLUMN_WIDTH)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Auto-filter: create a proper Excel Table object (supports multiple per sheet)
    if auto_filter:
        num_cols = len(table_data[0]) if table_data else 0
        if num_cols > 0:
            last_col_letter = get_column_letter(num_cols)
            last_data_row = start_row + len(table_data) - 1
            table_ref = f"A{start_row}:{last_col_letter}{last_data_row}"
            # Excel table names must be unique across the workbook
            table_name = f"Table_{worksheet.title.replace(' ', '_')}_{table_index + 1}"
            # Sanitize: Excel table names allow only letters, digits, underscores
            table_name = re.sub(r'[^A-Za-z0-9_]', '', table_name)
            excel_table = Table(displayName=table_name, ref=table_ref)
            excel_table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2", showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False,
            )
            worksheet.add_table(excel_table)

    return start_row + len(table_data) + TABLE_BOTTOM_SPACING
