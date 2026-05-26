"""Tests for Excel (xlsx) multi-sheet support.

These tests verify that the markdown to Excel conversion handles
the '## Sheet: Name' heading syntax correctly for multi-sheet workbooks.
"""

import inspect
import sys
from pathlib import Path
from unittest.mock import patch, MagicMock

# Add project root to path for imports
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

import pytest
from openpyxl import Workbook, load_workbook
import io

# We test the internal parsing logic, mocking the upload step.
from xlsx_tools.base_xlsx_tool import markdown_to_excel


def _create_workbook_from_markdown(markdown_content: str, save_name: str | None = None) -> Workbook:
    """Helper that runs markdown_to_excel but intercepts the workbook before upload.

    Patches upload_file to capture the BytesIO and returns a loaded Workbook.
    Automatically saves output to OUTPUT_DIR for manual inspection.
    """
    captured = {}

    def fake_upload(file_obj, suffix, **kwargs):
        captured['data'] = file_obj.read()
        file_obj.seek(0)
        return "https://fake-url/test.xlsx"

    with patch("xlsx_tools.base_xlsx_tool.upload_file", side_effect=fake_upload):
        markdown_to_excel(markdown_content)

    data = captured['data']
    wb = load_workbook(io.BytesIO(data))

    # Auto-derive save name from calling test function
    if save_name is None:
        frame = inspect.stack()[1]
        save_name = frame.function

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / f"{save_name}.xlsx"
    out_path.write_bytes(data)

    return wb


# Output directory for test files
OUTPUT_DIR = Path(__file__).parent / "output" / "xlsx"


@pytest.fixture(scope="module", autouse=True)
def setup_output_dir():
    """Create output directory if it doesn't exist."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    yield


class TestMultiSheet:
    """Tests for multi-sheet Excel workbooks via ## Sheet: Name."""

    def test_single_sheet_default_name(self):
        """Markdown without ## Sheet: heading → single sheet named 'Data Report'."""
        markdown = """# Report

| Name | Value |
|------|-------|
| A    | 1     |
"""
        wb = _create_workbook_from_markdown(markdown)
        assert len(wb.sheetnames) == 1
        assert wb.sheetnames[0] == "Data Report"

    def test_two_sheets(self):
        """Markdown with two ## Sheet: headings → two sheets."""
        markdown = """## Sheet: Revenue

| Quarter | Amount |
|---------|--------|
| Q1      | 1000   |
| Q2      | 1200   |

## Sheet: Expenses

| Quarter | Amount |
|---------|--------|
| Q1      | 800    |
| Q2      | 900    |
"""
        wb = _create_workbook_from_markdown(markdown)
        assert len(wb.sheetnames) == 2
        assert "Revenue" in wb.sheetnames
        assert "Expenses" in wb.sheetnames

    def test_sheet_names_correct(self):
        """Verify sheet names are correctly set from headings."""
        markdown = """## Sheet: Summary

| Metric | Value |
|--------|-------|
| Total  | 100   |

## Sheet: Detail Data

| Item | Count |
|------|-------|
| A    | 50    |
| B    | 50    |
"""
        wb = _create_workbook_from_markdown(markdown)
        assert wb.sheetnames[0] == "Summary"
        assert wb.sheetnames[1] == "Detail Data"

    def test_data_on_correct_sheets(self):
        """Verify tables land on the correct sheets."""
        markdown = """## Sheet: Sheet1

| Col1 | Col2 |
|------|------|
| X    | Y    |

## Sheet: Sheet2

| ColA | ColB |
|------|------|
| M    | N    |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws1 = wb["Sheet1"]
        ws2 = wb["Sheet2"]

        # Check data exists on Sheet1 (header row + data row)
        # The exact row depends on spacing; just verify the values exist somewhere
        sheet1_values = []
        for row in ws1.iter_rows(values_only=True):
            sheet1_values.extend([v for v in row if v is not None])
        assert "X" in sheet1_values or "x" in str(sheet1_values).lower()

        sheet2_values = []
        for row in ws2.iter_rows(values_only=True):
            sheet2_values.extend([v for v in row if v is not None])
        assert "M" in sheet2_values or "m" in str(sheet2_values).lower()

    def test_three_sheets(self):
        """Test creating three sheets."""
        markdown = """## Sheet: Alpha

| A |
|---|
| 1 |

## Sheet: Beta

| B |
|---|
| 2 |

## Sheet: Gamma

| C |
|---|
| 3 |
"""
        wb = _create_workbook_from_markdown(markdown)
        assert len(wb.sheetnames) == 3
        assert wb.sheetnames == ["Alpha", "Beta", "Gamma"]

    def test_backwards_compatible_no_sheet_heading(self):
        """Without any ## Sheet: headings, everything goes to 'Data Report'."""
        markdown = """# My Report

| Name | Age |
|------|-----|
| Alice | 30 |
| Bob   | 25 |
"""
        wb = _create_workbook_from_markdown(markdown)
        assert len(wb.sheetnames) == 1
        assert wb.sheetnames[0] == "Data Report"


class TestCrossSheetReferences:
    """Tests for cross-sheet cell references via SheetName!T1.B[0] syntax."""

    def test_simple_cross_sheet_reference(self):
        """A formula on Sheet2 referencing a cell on Sheet1 via SheetName!T1.B[0]."""
        markdown = """## Sheet: Revenue

| Quarter | Amount |
|---------|--------|
| Q1      | 1000   |
| Q2      | 1200   |

## Sheet: Dashboard

| Metric | Value |
|--------|-------|
| Q1 Rev | =Revenue!T1.B[0] |
| Q2 Rev | =Revenue!T1.B[1] |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb["Dashboard"]
        # Revenue table starts at row 1 (first sheet, no heading before table)
        # T1 on Revenue starts at row 1, data row [0] → row 2, data row [1] → row 3
        # So the formulas should be =Revenue!B2 and =Revenue!B3
        q1_cell = ws.cell(row=1, column=2)  # header row of Dashboard T1 is row 1
        q2_cell = ws.cell(row=2, column=2)  # data row 0
        # Data rows of Dashboard are at row 2, row 3
        q1_val = ws.cell(row=2, column=2).value
        q2_val = ws.cell(row=3, column=2).value
        assert q1_val == "=Revenue!B2", f"Expected =Revenue!B2, got {q1_val}"
        assert q2_val == "=Revenue!B3", f"Expected =Revenue!B3, got {q2_val}"

    def test_cross_sheet_reference_with_spaces_in_name(self):
        """Sheet names with spaces get quoted in the Excel formula."""
        markdown = """## Sheet: Sales Data

| Product | Revenue |
|---------|---------|
| Widget  | 5000    |

## Sheet: Summary

| Metric | Value |
|--------|-------|
| Total  | =Sales Data!T1.B[0] |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb["Summary"]
        cell_val = ws.cell(row=2, column=2).value
        assert cell_val == "='Sales Data'!B2", f"Expected ='Sales Data'!B2, got {cell_val}"

    def test_forward_reference_sheet1_refs_sheet2(self):
        """Sheet1 formula references Sheet2 that hasn't been parsed yet (forward reference)."""
        markdown = """## Sheet: Dashboard

| Metric | Value |
|--------|-------|
| Total  | =Details!T1.B[0] |

## Sheet: Details

| Item | Amount |
|------|--------|
| Rent | 3000   |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb["Dashboard"]
        cell_val = ws.cell(row=2, column=2).value
        assert cell_val == "=Details!B2", f"Expected =Details!B2, got {cell_val}"

    def test_cross_sheet_range_reference(self):
        """Cross-sheet range reference SheetName!T1.B[0]:T1.B[2]."""
        markdown = """## Sheet: Data

| Name | Score |
|------|-------|
| A    | 10    |
| B    | 20    |
| C    | 30    |

## Sheet: Summary

| Metric | Value |
|--------|-------|
| Total  | =SUM(Data!T1.B[0]:T1.B[2]) |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb["Summary"]
        cell_val = ws.cell(row=2, column=2).value
        assert cell_val == "=SUM(Data!B2:B4)", f"Expected =SUM(Data!B2:B4), got {cell_val}"

    def test_mixed_local_and_cross_sheet(self):
        """Formula mixing local and cross-sheet references."""
        markdown = """## Sheet: Revenue

| Quarter | Amount |
|---------|--------|
| Q1      | 1000   |

## Sheet: Costs

| Quarter | Amount |
|---------|--------|
| Q1      | 400    |

## Sheet: Profit

| Quarter | Revenue | Cost | Profit |
|---------|---------|------|--------|
| Q1      | =Revenue!T1.B[0] | =Costs!T1.B[0] | =B[0]-C[0] |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb["Profit"]
        rev_val = ws.cell(row=2, column=2).value
        cost_val = ws.cell(row=2, column=3).value
        profit_val = ws.cell(row=2, column=4).value
        assert rev_val == "=Revenue!B2", f"Expected =Revenue!B2, got {rev_val}"
        assert cost_val == "=Costs!B2", f"Expected =Costs!B2, got {cost_val}"
        assert profit_val == "=B2-C2", f"Expected =B2-C2, got {profit_val}"

    def test_cross_sheet_with_header_before_table(self):
        """Cross-sheet reference where the source sheet has a header before the table."""
        markdown = """## Sheet: Source

# Monthly Data

| Month | Value |
|-------|-------|
| Jan   | 100   |
| Feb   | 200   |

## Sheet: Target

| Metric | Value |
|--------|-------|
| Jan    | =Source!T1.B[0] |
| Feb    | =Source!T1.B[1] |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb["Target"]
        # Source: header at row 1, +2 spacing → table starts at row 3
        # T1 data[0] = row 4, data[1] = row 5
        jan_val = ws.cell(row=2, column=2).value
        feb_val = ws.cell(row=3, column=2).value
        assert jan_val == "=Source!B4", f"Expected =Source!B4, got {jan_val}"
        assert feb_val == "=Source!B5", f"Expected =Source!B5, got {feb_val}"


class TestAdjustFormulaReferencesUnit:
    """Unit tests for adjust_formula_references with cross-sheet support."""

    def test_cross_sheet_single_cell(self):
        from xlsx_tools.helpers import adjust_formula_references
        all_positions = {"Revenue": {"T1": 1}}
        result = adjust_formula_references(
            "=Revenue!T1.B[0]", 10, {}, all_positions
        )
        assert result == "=Revenue!B2"

    def test_cross_sheet_quoted_name(self):
        from xlsx_tools.helpers import adjust_formula_references
        all_positions = {"My Sheet": {"T1": 1}}
        result = adjust_formula_references(
            "=My Sheet!T1.A[2]", 10, {}, all_positions
        )
        assert result == "='My Sheet'!A4"

    def test_cross_sheet_range(self):
        from xlsx_tools.helpers import adjust_formula_references
        all_positions = {"Data": {"T1": 1}}
        result = adjust_formula_references(
            "=SUM(Data!T1.B[0]:T1.B[4])", 10, {}, all_positions
        )
        assert result == "=SUM(Data!B2:B6)"

    def test_cross_sheet_function_pattern(self):
        from xlsx_tools.helpers import adjust_formula_references
        all_positions = {"Sales": {"T1": 3}}
        result = adjust_formula_references(
            "=Sales!T1.SUM(B[0]:D[0])", 10, {}, all_positions
        )
        # T1 starts at row 3, data[0] → row 4
        assert result == "=SUM(Sales!B4:Sales!D4)"

    def test_local_reference_still_works(self):
        from xlsx_tools.helpers import adjust_formula_references
        result = adjust_formula_references(
            "=T1.B[0]", 5, {"T1": 1}, {}
        )
        assert result == "=B2"

    def test_mixed_local_and_cross_sheet(self):
        from xlsx_tools.helpers import adjust_formula_references
        all_positions = {"Revenue": {"T1": 1}}
        result = adjust_formula_references(
            "=Revenue!T1.B[0]-B[0]", 5, {"T1": 3}, all_positions
        )
        # Revenue!T1.B[0] → Revenue!B2, B[0] → B4 (current table starts at 3, data[0] = row 4)
        assert result == "=Revenue!B2-B4"


class TestNumberFormats:
    """Tests for cell number_format (percent, thousands separator)."""

    def test_percent_cells_get_percent_format(self):
        """Cells with '50%' should be stored as 0.5 with number_format '0%'."""
        markdown = """| Metric | Rate |
|--------|------|
| Growth | 50%  |
| Margin | 8%   |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        # Data rows start at row 2 (row 1 = header)
        growth_cell = ws.cell(row=2, column=2)
        margin_cell = ws.cell(row=3, column=2)
        assert growth_cell.value == pytest.approx(0.5)
        assert growth_cell.number_format == '0%'
        assert margin_cell.value == pytest.approx(0.08)
        assert margin_cell.number_format == '0%'

    def test_non_percent_number_no_percent_format(self):
        """A plain decimal like 0.5 (without '%') should NOT get '0%' format."""
        markdown = """| Item  | Value |
|-------|-------|
| Alpha | 0.5   |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        cell = ws.cell(row=2, column=2)
        assert cell.value == pytest.approx(0.5)
        assert cell.number_format != '0%'

    def test_thousands_separator_format(self):
        """Values >= 1000 should get '#,##0' number format."""
        markdown = """| Item | Amount |
|------|--------|
| Rent | 5000   |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        cell = ws.cell(row=2, column=2)
        assert cell.value == 5000
        assert cell.number_format == '#,##0'


class TestAutoFilter:
    """Tests for the auto_filter feature."""

    def test_auto_filter_enabled(self):
        """When auto_filter=True, auto-filter is applied to the table range."""
        markdown = """| Name | Value |
|------|-------|
| A    | 1     |
| B    | 2     |
"""
        captured = {}

        def fake_upload(file_obj, suffix, **kwargs):
            captured['data'] = file_obj.read()
            file_obj.seek(0)
            return "https://fake-url/test.xlsx"

        with patch("xlsx_tools.base_xlsx_tool.upload_file", side_effect=fake_upload):
            markdown_to_excel(markdown, auto_filter=True)

        wb = load_workbook(io.BytesIO(captured['data']))
        ws = wb.active
        # Now uses Excel Table objects (each table has its own filter)
        assert len(ws.tables) == 1
        table = list(ws.tables.values())[0]
        assert table.ref == "A1:B3"

    def test_auto_filter_disabled_by_default(self):
        """By default, no auto-filter is applied."""
        markdown = """| Name | Value |
|------|-------|
| A    | 1     |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert len(ws.tables) == 0

    def test_auto_filter_multi_sheet(self):
        """Auto-filter works on multi-sheet workbooks (applied per table)."""
        markdown = """## Sheet: Data

| Col1 | Col2 | Col3 |
|------|------|------|
| A    | B    | C    |
| D    | E    | F    |
"""
        captured = {}

        def fake_upload(file_obj, suffix, **kwargs):
            captured['data'] = file_obj.read()
            file_obj.seek(0)
            return "https://fake-url/test.xlsx"

        with patch("xlsx_tools.base_xlsx_tool.upload_file", side_effect=fake_upload):
            markdown_to_excel(markdown, auto_filter=True)

        wb = load_workbook(io.BytesIO(captured['data']))
        ws = wb["Data"]
        assert len(ws.tables) == 1
        table = list(ws.tables.values())[0]
        assert table.ref == "A1:C3"

    def test_auto_filter_numeric_header_stays_string(self):
        """Headers that look like numbers remain strings (Excel Table requires string headers)."""
        markdown = """| 2024 | 2025 | Growth |
|------|------|--------|
| 100  | 150  | 50%    |
"""
        captured = {}

        def fake_upload(file_obj, suffix, **kwargs):
            captured['data'] = file_obj.read()
            file_obj.seek(0)
            return "https://fake-url/test.xlsx"

        with patch("xlsx_tools.base_xlsx_tool.upload_file", side_effect=fake_upload):
            markdown_to_excel(markdown, auto_filter=True)

        wb = load_workbook(io.BytesIO(captured['data']))
        ws = wb.active
        # Headers must be strings, not numbers
        assert ws.cell(row=1, column=1).value == "2024"
        assert isinstance(ws.cell(row=1, column=1).value, str)
        assert ws.cell(row=1, column=2).value == "2025"
        assert isinstance(ws.cell(row=1, column=2).value, str)


class TestColumnAlignment:
    """Tests for column alignment from separator row markers."""

    def test_explicit_alignment_left_center_right(self):
        """Separator markers :--- / :---: / ---: set column alignment."""
        markdown = """| Name | Status | Amount |
|:-----|:------:|-------:|
| Alice | Active | 1000   |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        # Data row (row 2): check alignment
        assert ws.cell(row=2, column=1).alignment.horizontal == 'left'
        assert ws.cell(row=2, column=2).alignment.horizontal == 'center'
        assert ws.cell(row=2, column=3).alignment.horizontal == 'right'

    def test_no_alignment_markers_uses_heuristic(self):
        """Without alignment markers, numbers align right and text aligns left."""
        markdown = """| Name | Value |
|------|-------|
| Alice | 100  |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.cell(row=2, column=1).alignment.horizontal == 'left'
        assert ws.cell(row=2, column=2).alignment.horizontal == 'right'

    def test_header_row_always_centered(self):
        """Header row is always centered regardless of alignment markers."""
        markdown = """| Name | Value |
|:-----|------:|
| A    | 1     |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.cell(row=1, column=1).alignment.horizontal == 'center'
        assert ws.cell(row=1, column=2).alignment.horizontal == 'center'


class TestDateDetection:
    """Tests for automatic date detection and formatting."""

    def test_iso_date(self):
        """ISO dates (YYYY-MM-DD) are detected and stored as datetime."""
        markdown = """| Event | Date |
|-------|------|
| Launch | 2024-01-15 |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        cell = ws.cell(row=2, column=2)
        from datetime import datetime
        assert isinstance(cell.value, datetime)
        assert cell.value.year == 2024
        assert cell.value.month == 1
        assert cell.value.day == 15
        assert cell.number_format == "YYYY-MM-DD"

    def test_european_date_dot_format(self):
        """European dates (DD.MM.YYYY) are detected."""
        markdown = """| Date |
|------|
| 15.03.2024 |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        cell = ws.cell(row=2, column=1)
        from datetime import datetime
        assert isinstance(cell.value, datetime)
        assert cell.value.day == 15
        assert cell.value.month == 3
        assert cell.number_format == "DD.MM.YYYY"

    def test_named_month_date(self):
        """Named month dates (Jan 15, 2024) are detected."""
        markdown = """| Date |
|------|
| Jan 15, 2024 |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        cell = ws.cell(row=2, column=1)
        from datetime import datetime
        assert isinstance(cell.value, datetime)
        assert cell.value.month == 1
        assert cell.value.day == 15

    def test_iso_datetime_with_time(self):
        """ISO datetime with time component is detected."""
        markdown = """| Timestamp |
|-----------|
| 2024-01-15T14:30 |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        cell = ws.cell(row=2, column=1)
        from datetime import datetime
        assert isinstance(cell.value, datetime)
        assert cell.value.hour == 14
        assert cell.value.minute == 30
        assert cell.number_format == "YYYY-MM-DD HH:MM"

    def test_plain_numbers_not_detected_as_dates(self):
        """Plain numbers should NOT be detected as dates."""
        markdown = """| Value |
|-------|
| 1234  |
| 2024  |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == 1234.0
        assert ws.cell(row=3, column=1).value == 2024.0

    def test_short_strings_not_detected_as_dates(self):
        """Short strings like 'Q1', 'N/A' should NOT be detected as dates."""
        markdown = """| Label |
|-------|
| Q1    |
| N/A   |
| Hello |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "Q1"
        assert ws.cell(row=3, column=1).value == "N/A"
        assert ws.cell(row=4, column=1).value == "Hello"


class TestDateDetectionUnit:
    """Unit tests for _try_parse_date."""

    def test_iso_format(self):
        from xlsx_tools.helpers import _try_parse_date
        result = _try_parse_date("2024-06-15")
        assert result is not None
        dt, fmt = result
        assert dt.year == 2024 and dt.month == 6 and dt.day == 15
        assert fmt == "YYYY-MM-DD"

    def test_european_dot(self):
        from xlsx_tools.helpers import _try_parse_date
        result = _try_parse_date("31.12.2023")
        assert result is not None
        dt, fmt = result
        assert dt.day == 31 and dt.month == 12 and dt.year == 2023
        assert fmt == "DD.MM.YYYY"

    def test_short_year(self):
        from xlsx_tools.helpers import _try_parse_date
        result = _try_parse_date("15.03.24")
        assert result is not None
        dt, fmt = result
        assert dt.day == 15 and dt.month == 3
        assert fmt == "DD.MM.YY"

    def test_named_month_long(self):
        from xlsx_tools.helpers import _try_parse_date
        result = _try_parse_date("15 March 2024")
        assert result is not None
        dt, fmt = result
        assert dt.day == 15 and dt.month == 3
        assert fmt == "DD MMMM YYYY"

    def test_rejects_plain_number(self):
        from xlsx_tools.helpers import _try_parse_date
        assert _try_parse_date("12345") is None
        assert _try_parse_date("2024") is None

    def test_rejects_short_string(self):
        from xlsx_tools.helpers import _try_parse_date
        assert _try_parse_date("Q1") is None
        assert _try_parse_date("Hello") is None
        assert _try_parse_date("AB") is None

    def test_rejects_empty(self):
        from xlsx_tools.helpers import _try_parse_date
        assert _try_parse_date("") is None
        assert _try_parse_date("abc") is None


class TestTableParsingImprovements:
    """Tests for parse_table improvements (trailing pipe, separator detection)."""

    def test_table_without_trailing_pipe(self):
        """Tables without trailing pipes are parsed correctly."""
        markdown = """| Name | Value
|------|------
| A    | 1
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "A"

    def test_data_with_dashes_not_skipped(self):
        """Data rows containing '---' are NOT skipped as separator rows."""
        markdown = """| Status | Code |
|--------|------|
| error---retry | 500 |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "error---retry"
        assert ws.cell(row=2, column=2).value == 500.0


class TestFreezeDirective:
    """Tests for <!-- freeze --> directive."""

    def test_freeze_panes_applied(self):
        """Freeze directive freezes below the table header row."""
        markdown = """<!-- freeze -->
| Name | Value |
|------|-------|
| A    | 1     |
| B    | 2     |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.freeze_panes == "A2"

    def test_freeze_not_applied_by_default(self):
        """Without freeze directive, no freeze panes."""
        markdown = """| Name | Value |
|------|-------|
| A    | 1     |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.freeze_panes is None

    def test_freeze_with_header_offset(self):
        """Freeze works when table starts after a header."""
        markdown = """# Title

<!-- freeze -->
| Name | Value |
|------|-------|
| A    | 1     |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        # Table starts at row 3 (after header at row 1 + spacing), so freeze at A4
        assert ws.freeze_panes == "A4"


class TestTypesDirective:
    """Tests for <!-- types: ... --> directive."""

    def test_currency_type(self):
        """Currency directive strips symbol and stores as number."""
        markdown = """<!-- types: text, currency:$ -->
| Item  | Price    |
|-------|----------|
| Apple | $1,234.56 |
| Banana | $50.00  |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.cell(row=2, column=2).value == 1234.56
        assert ws.cell(row=3, column=2).value == 50.00
        assert '$' in ws.cell(row=2, column=2).number_format

    def test_bool_type(self):
        """Bool directive converts true/false strings to Excel booleans."""
        markdown = """<!-- types: text, bool -->
| Name | Active |
|------|--------|
| A    | true   |
| B    | false  |
| C    | yes    |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.cell(row=2, column=2).value is True
        assert ws.cell(row=3, column=2).value is False
        assert ws.cell(row=4, column=2).value is True

    def test_text_type_prevents_conversion(self):
        """Text directive keeps values as strings even if they look like numbers."""
        markdown = """<!-- types: text -->
| Code  |
|-------|
| 00123 |
| 45678 |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "00123"
        assert ws.cell(row=3, column=1).value == "45678"

    def test_euro_currency(self):
        """European currency with dot-as-thousands separator."""
        markdown = """<!-- types: currency:€ -->
| Amount     |
|------------|
| €1.234,56  |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert abs(ws.cell(row=2, column=1).value - 1234.56) < 0.01
        assert '€' in ws.cell(row=2, column=1).number_format

    def test_date_type_with_format(self):
        """Date type with explicit format."""
        markdown = """<!-- types: text, date:DD.MM.YYYY -->
| Name | Born       |
|------|------------|
| Alice | 15.03.1990 |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        cell_val = ws.cell(row=2, column=2).value
        assert hasattr(cell_val, 'year')  # It's a datetime
        assert cell_val.day == 15
        assert cell_val.month == 3
        assert ws.cell(row=2, column=2).number_format == "DD.MM.YYYY"

    def test_partial_types_only_some_columns(self):
        """Types directive can specify fewer types than columns — rest use auto-detection."""
        markdown = """<!-- types: text, number -->
| Code | Amount | Date       |
|------|--------|------------|
| 007  | 1500   | 2024-01-15 |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        # First col: forced text
        assert ws.cell(row=2, column=1).value == "007"
        # Second col: forced number
        assert ws.cell(row=2, column=2).value == 1500.0
        # Third col: no type spec → auto-detection (date)
        cell_val = ws.cell(row=2, column=3).value
        assert hasattr(cell_val, 'year')


class TestTypesDirectiveAdvanced:
    """Advanced tests for <!-- types: ... --> directive edge cases and interactions."""

    def test_percent_type(self):
        """Percent directive converts '75%' to 0.75 with 0% format."""
        markdown = """<!-- types: text, percent -->
| Task | Progress |
|------|----------|
| A    | 75%      |
| B    | 100%     |
| C    | 0%       |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.cell(row=2, column=2).value == pytest.approx(0.75)
        assert ws.cell(row=2, column=2).number_format == '0%'
        assert ws.cell(row=3, column=2).value == pytest.approx(1.0)
        assert ws.cell(row=4, column=2).value == pytest.approx(0.0)

    def test_number_type_with_format(self):
        """Number directive with explicit format applies the format."""
        markdown = """<!-- types: number:0.00 -->
| Score |
|-------|
| 3.14159 |
| 2.71828 |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == pytest.approx(3.14159)
        assert ws.cell(row=2, column=1).number_format == '0.00'

    def test_number_type_without_format(self):
        """Number directive without format applies #,##0 for large numbers."""
        markdown = """<!-- types: number -->
| Amount |
|--------|
| 50000  |
| 500    |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == 50000.0
        assert ws.cell(row=2, column=1).number_format == '#,##0'
        # Small number — no forced format
        assert ws.cell(row=3, column=1).value == 500.0

    def test_unknown_type_falls_through(self):
        """An unknown type spec falls through to normal auto-detection."""
        markdown = """<!-- types: unknowntype -->
| Value |
|-------|
| 1234  |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        # Falls through to normal resolve_cell, which detects a number
        assert ws.cell(row=2, column=1).value == 1234.0

    def test_types_only_apply_to_data_rows(self):
        """Type directives do NOT affect the header row."""
        markdown = """<!-- types: number -->
| Amount |
|--------|
| 100    |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        # Header row should still be text "Amount" with header styling
        assert ws.cell(row=1, column=1).value == "Amount"
        assert ws.cell(row=1, column=1).font.bold is True

    def test_types_with_alignment(self):
        """Types directive respects explicit column alignment from separator row."""
        markdown = """<!-- types: text, currency:$ -->
| Item  | Price    |
|:------|:--------:|
| Apple | $50.00   |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        # Left alignment from separator
        assert ws.cell(row=2, column=1).alignment.horizontal == 'left'
        # Center alignment from separator overrides the default 'right' for numbers
        assert ws.cell(row=2, column=2).alignment.horizontal == 'center'

    def test_bool_alignment_center(self):
        """Bool values get center alignment when no explicit alignment specified."""
        markdown = """<!-- types: bool -->
| Active |
|--------|
| true   |
| false  |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.cell(row=2, column=1).value is True
        assert ws.cell(row=2, column=1).alignment.horizontal == 'center'
        assert ws.cell(row=3, column=1).value is False
        assert ws.cell(row=3, column=1).alignment.horizontal == 'center'

    def test_currency_czk(self):
        """Czech koruna currency format."""
        markdown = """<!-- types: currency:Kč -->
| Cena     |
|----------|
| 1 234,56 Kč |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert abs(ws.cell(row=2, column=1).value - 1234.56) < 0.01
        assert 'Kč' in ws.cell(row=2, column=1).number_format

    def test_multiple_directives_stacked(self):
        """Multiple directives above a table all take effect."""
        markdown = """<!-- freeze -->
<!-- types: text, number -->
| Code | Value |
|------|-------|
| ABC  | 5000  |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        # Freeze should be applied
        assert ws.freeze_panes == "A2"
        # Types should be applied
        assert ws.cell(row=2, column=1).value == "ABC"
        assert ws.cell(row=2, column=2).value == 5000.0

    def test_formatted_text_with_type_directive(self):
        """Inline markdown formatting is stripped before type coercion and applied to cell."""
        markdown = """<!-- types: number, currency:$ -->
| Quantity | Price  |
|----------|--------|
| **123**  | *50.00* |
| `456`    | 100    |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        # Bold 123 should be parsed as number with bold formatting
        assert ws.cell(row=2, column=1).value == 123.0
        assert ws.cell(row=2, column=1).font.bold is True
        # Italic 50.00 should be parsed as currency with italic formatting
        assert ws.cell(row=2, column=2).value == 50.0
        assert ws.cell(row=2, column=2).font.italic is True
        # Monospace 456 should be parsed as number with Courier New
        assert ws.cell(row=3, column=1).value == 456.0
        assert ws.cell(row=3, column=1).font.name == 'Courier New'

    def test_directives_dont_carry_across_tables(self):
        """Directives only apply to the immediately following table."""
        markdown = """<!-- types: text -->
| Code |
|------|
| 007  |

| Amount |
|--------|
| 1234   |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        # First table: text directive forces string
        assert ws.cell(row=2, column=1).value == "007"
        # Second table: no directive, auto-detection → number
        # Second table starts at row 1 + 2 (table size) + 2 (spacing) = row 5
        # Actually need to find it dynamically
        found_number = False
        for row in ws.iter_rows(min_row=3, values_only=True):
            for val in row:
                if val == 1234.0:
                    found_number = True
        assert found_number, "Second table should auto-detect 1234 as a number"

    def test_directives_dont_carry_across_prose(self):
        """Directives are cleared if non-table content appears between directive and table."""
        markdown = """<!-- types: text -->

Some prose paragraph here.

| Amount |
|--------|
| 1234   |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        # The directive should NOT apply because prose intervened
        found_number = False
        for row in ws.iter_rows(values_only=True):
            for val in row:
                if val == 1234.0:
                    found_number = True
        assert found_number, "Directive should not apply when prose separates it from the table"

    def test_directives_dont_carry_across_sheets(self):
        """Directives are reset when a new sheet is started."""
        markdown = """## Sheet: Sheet1

<!-- types: text -->
| Code |
|------|
| 007  |

## Sheet: Sheet2

| Code |
|------|
| 007  |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws1 = wb["Sheet1"]
        ws2 = wb["Sheet2"]
        # Sheet1: text directive forces string
        assert ws1.cell(row=2, column=1).value == "007"
        # Sheet2: no directive → auto-detection would parse as number
        # (007 → 7.0 without text directive)
        assert ws2.cell(row=2, column=1).value == 7.0

    def test_date_type_without_format(self):
        """Date type without explicit format uses auto-detected format."""
        markdown = """<!-- types: date -->
| When       |
|------------|
| 2024-06-15 |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        cell = ws.cell(row=2, column=1)
        assert hasattr(cell.value, 'year')
        assert cell.value.year == 2024
        assert cell.value.month == 6
        assert cell.value.day == 15

    def test_currency_unparseable_value(self):
        """Currency type with unparseable value falls back to text."""
        markdown = """<!-- types: currency:$ -->
| Price   |
|---------|
| N/A     |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "N/A"

    def test_currency_empty_symbol_defaults_to_dollar(self):
        """Currency directive with no symbol (currency:) defaults to $ without raising."""
        markdown = """<!-- types: currency: -->
| Price   |
|---------|
| 1234.56 |
"""
        wb = _create_workbook_from_markdown(markdown)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == pytest.approx(1234.56)
        assert '$' in ws.cell(row=2, column=1).number_format


class TestDirectiveParserUnit:
    """Unit tests for _parse_types_directive helper."""

    def test_empty_string(self):
        from xlsx_tools.helpers import _parse_types_directive
        assert _parse_types_directive("") == []

    def test_single_type(self):
        from xlsx_tools.helpers import _parse_types_directive
        result = _parse_types_directive("text")
        assert result == ["text"]

    def test_multiple_types(self):
        from xlsx_tools.helpers import _parse_types_directive
        result = _parse_types_directive("text, currency:$, date, bool, number")
        assert result == ["text", "currency:$", "date", "bool", "number"]

    def test_blank_entries_become_none(self):
        from xlsx_tools.helpers import _parse_types_directive
        result = _parse_types_directive("text, , number")
        assert result == ["text", None, "number"]

    def test_whitespace_handling(self):
        from xlsx_tools.helpers import _parse_types_directive
        result = _parse_types_directive("  text  ,  number:0.00  ")
        assert result == ["text", "number:0.00"]


class TestApplyColumnTypeUnit:
    """Unit tests for _apply_column_type helper."""

    def test_text_type(self):
        from xlsx_tools.helpers import _apply_column_type
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        result = _apply_column_type(cell, "00123", "text")
        assert result is True
        assert cell.value == "00123"

    def test_bool_true_values(self):
        from xlsx_tools.helpers import _apply_column_type
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for val in ("true", "True", "yes", "YES", "1", "on"):
            cell = ws.cell(row=1, column=1)
            result = _apply_column_type(cell, val, "bool")
            assert result is True
            assert cell.value is True, f"Expected True for '{val}'"

    def test_bool_false_values(self):
        from xlsx_tools.helpers import _apply_column_type
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for val in ("false", "False", "no", "NO", "0", "off"):
            cell = ws.cell(row=1, column=1)
            result = _apply_column_type(cell, val, "bool")
            assert result is True
            assert cell.value is False, f"Expected False for '{val}'"

    def test_bool_unrecognized(self):
        from xlsx_tools.helpers import _apply_column_type
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        result = _apply_column_type(cell, "maybe", "bool")
        assert result is True
        assert cell.value == "maybe"

    def test_percent(self):
        from xlsx_tools.helpers import _apply_column_type
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        result = _apply_column_type(cell, "45%", "percent")
        assert result is True
        assert cell.value == pytest.approx(0.45)
        assert cell.number_format == '0%'

    def test_none_type_returns_false(self):
        from xlsx_tools.helpers import _apply_column_type
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        result = _apply_column_type(cell, "hello", None)
        assert result is False

    def test_unknown_type_returns_false(self):
        from xlsx_tools.helpers import _apply_column_type
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1)
        result = _apply_column_type(cell, "hello", "foobar")
        assert result is False


if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])

